# -*- coding: utf-8 -*-

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import Firefox, FirefoxProfile
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.proxy import *

import os
import re
import requests
from datetime import datetime
import time
import pause
import random
from openpyxl import load_workbook
import json
import traceback
import math
import threading
import asyncio

filePath = os.path.dirname(os.path.abspath(__file__))

class nikeBot:
    def __init__(self, whichtask, sheet):
        self.whichtask = whichtask
        self.sheet = sheet
        self.login = (str(sheet.cell(row=whichtask, column=1).value)).replace(' ', '')
        self.password = (str(sheet.cell(row=whichtask, column=2).value)).replace(' ', '')
        self.name = str(sheet.cell(row=1, column=2).value)
        diapason = int(math.log2(random.randint(1, 1023)))
        for _ in range(0, diapason):
            self.name += ',' if random.randint(1, 2) == 1 else '.'

        self.surname = str(sheet.cell(row=2, column=2).value)
        diapason = int(math.log2(random.randint(1, 1023)))
        for _ in range(0, diapason):
            self.surname += ',' if random.randint(1, 2) == 1 else '.'

        self.middlename = str(sheet.cell(row=3, column=2).value)
        diapason = int(math.log2(random.randint(1, 1023)))
        for _ in range(0, diapason):
            self.middlename += ',' if random.randint(1, 2) == 1 else '.'

        self.addressline = str(sheet.cell(row=4, column=2).value)
        diapason = int(math.log2(random.randint(1, 1023)))
        for _ in range(0, diapason):
            self.addressline += ',' if random.randint(1, 2) == 1 else '.'
        
        self.city = str(sheet.cell(row=5, column=2).value)
        self.index = (str(sheet.cell(row=6, column=2).value)).replace(' ', '')
        self.phone = (str(sheet.cell(row=whichtask, column=3).value)).replace(' ', '')
        self.email = (str(sheet.cell(row=whichtask, column=4).value)).replace(' ', '')
        self.cardcvc = (str(sheet.cell(row=whichtask, column=5).value)).replace(' ', '')
        self.releaselink = (str(sheet.cell(row=whichtask, column=8).value)).replace(' ', '')
        self.sec = (str(sheet.cell(row=whichtask, column=6).value)).replace(' ', '')
        self.minute = (str(sheet.cell(row=2, column=8).value)).replace(' ', '')
        self.hour = (str(sheet.cell(row=2, column=7).value)).replace(' ', '')
        self.price = (str(sheet.cell(row=2, column=6).value)).replace(' ', '')
  
        #proxy
        s = (str(sheet.cell(row=whichtask, column=7).value)).replace(' ', '')
        ip = s[:s.find(':')]
        s = s[s.find(':')+1:]
        port = s[:s.find(':')]
        s = s[s.find(':')+1:]
        proxyLogin = s[:s.find(':')]
        proxyPassword = s[s.rfind(':')+1:]

        #driver & requests
        self.session = requests.Session()
        
        if len(ip) > 5:
            proxy = Proxy({
                'proxyType': ProxyType.MANUAL,
                'httpProxy': ip + ":" + port,
                'socksUsername': proxyLogin,
                'socksPassword': proxyPassword
                })
            
            self.driver = Firefox(executable_path=filePath + '/geckodriver.exe', proxy=proxy)
            self.session.proxies.update({'https' : 'https://' + proxyLogin + ':' + proxyPassword + '@' + ip + ':' + port} )
        else:
            self.driver = Firefox(executable_path=filePath + '/geckodriver.exe')

    async def main(self):
        try:
            check = 0
            while check <= 2:
                try:
                    await self.logging()
                    await self.get_auth()
                    check = 3
                except: errors('in login ' + traceback.format_exc()); check += 1; await asyncio.sleep(6) 
            
            await asyncio.sleep(3)

            check = 0
            while check <= 2:
                try:
                    await self.checkout()
                    check = 3
                except: errors('in atk ' + traceback.format_exc()); check += 1; await asyncio.sleep(6) 
            
            try:
                await self.button_continue()
                await asyncio.sleep(1)
                await self.lastCard()
                await asyncio.sleep(1)
                await self.button_continue()
            except: errors('in card ' + traceback.format_exc())
            
            await self.atk_request()
            
            await asyncio.sleep(7200)
            self.driver.quit()
        except: 
            errors('in main ' + traceback.format_exc())
            self.driver.quit()
                
    async def get_auth(self):
        self.driver.get('https://unite.nike.com/auth/unite_session_cookies/v1')
        await asyncio.sleep(5)
        response = self.driver.find_element_by_xpath('//tr[@id="/access_token"]/td[2]/span/span')
        
        self.auth = response.text
        self.auth = self.auth[1:-1]

    async def logging(self):
        try:
            url = 'https://www.nike.com/ru/t/%D0%BA%D1%80%D0%BE%D1%81%D1%81%D0%BE%D0%B2%D0%BA%D0%B8-air-force-1-07-KyTDGepj/315122-111'
            self.driver.get(url)
            try: WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//button[@data-sub-type="image"]')))
            except: await asyncio.sleep(7)
            await asyncio.sleep(2)
            self.driver.find_element_by_xpath('//button[@data-sub-type="image"]').click()
            await asyncio.sleep(2)
            self.driver.find_element_by_xpath('//button[@class="g72-x-thick p10-sm text-color-primary-dark z1 css-1ddoumx"]').click()
            await asyncio.sleep(2)
        except: pass
        
        url = 'https://www.nike.com/ru/login'
        self.driver.get(url)
        try: WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@name="emailAddress"]')))
        except: await asyncio.sleep(7)
        email_form = self.driver.find_element_by_xpath('//input[@name="emailAddress"]')
        email_form.click()
        email_form.send_keys(self.login)
        
        await asyncio.sleep(1)
        password_form = self.driver.find_element_by_xpath('//input[@name="password"]')
        password_form.click()
        password_form.send_keys(self.password)
        
        await asyncio.sleep(1)
        self.driver.find_element_by_xpath('//input[@value="ВОЙТИ"]').click()
        await asyncio.sleep(6)
        self.driver.refresh()

    async def lastCard(self):
        iframe_el = self.driver.find_element_by_xpath('/html/body/esw-root/div/section/esw-checkout/div/esw-payment-details/div/div[2]/div[1]/esw-stored-cards-list/div/esw-stored-card/div/div/div[4]/esw-payment-iframe/div/iframe')
        self.driver.switch_to.frame(iframe_el)         
        
        cardcvc = self.cardcvc
        if len(cardcvc) > 3:
            cardcvc = cardcvc[:3]
        cvv = self.driver.find_element_by_xpath('//input[@id="cardCvc-input"]')
        cvv.clear()
        cvv.send_keys(cardcvc)

        self.driver.switch_to.default_content()
        await asyncio.sleep(1)

    async def button_continue(self):
        xpath = '//button[@class="button-continue"]'
        await self.wait_until_clickable((self.driver), xpath=xpath, duration=10)
        self.driver.find_element_by_xpath(xpath).click()
    
    async def checkout(self):
        self.driver.get(self.releaselink)
        await asyncio.sleep(7)
        try: WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//button[@class="button-continue"]')))
        except: await asyncio.sleep(7)
        await asyncio.sleep(3)

        await asyncio.sleep(1)
        name = self.driver.find_element_by_xpath('//input[@id="firstName"]')
        name.click()
        name.clear()
        if self.name != None: name.send_keys(self.name)

        await asyncio.sleep(1)
        lastname = self.driver.find_element_by_xpath('//input[@id="lastName"]')
        lastname.click()
        lastname.clear()
        if self.surname != None: lastname.send_keys(self.surname)
        
        await asyncio.sleep(1)
        middlename = self.driver.find_element_by_xpath('//input[@id="middleName"]')
        middlename.click()
        middlename.clear()
        if self.middlename != None: middlename.send_keys(self.middlename)
        
        await asyncio.sleep(1)
        address = self.driver.find_element_by_xpath('//input[@id="addressLine1"]')
        address.click()
        address.clear()
        if self.addressline != None: address.send_keys(self.addressline)
        
        await asyncio.sleep(1)
        address2 = self.driver.find_element_by_xpath('//input[@id="addressLine2"]')
        address2.click()
        address2.clear()
        address2.send_keys(" ")
        
        await asyncio.sleep(1)
        city = self.driver.find_element_by_xpath('//input[@id="city"]')
        city.click()
        await asyncio.sleep(1)
        city.clear()
        city.send_keys(self.city)
        actions = ActionChains(self.driver) 
        actions.send_keys(Keys.TAB * 2)
        actions.perform()
        
        await asyncio.sleep(2)
        phone = self.driver.find_element_by_xpath('//input[@placeholder="Номер телефона"]')
        phone.click()
        phone.clear()
        if self.phone != None: phone.send_keys(self.phone)
        
        await asyncio.sleep(1)
        index = self.driver.find_element_by_xpath('//input[@placeholder="Почтовый индекс"]')
        index.click()
        index.clear()
        if self.index != None: index.send_keys(self.index)

        await asyncio.sleep(1)
        email = self.driver.find_element_by_xpath('//input[@placeholder="Эл. почта"]')
        email.click()
        email.clear()
        if self.email != None: email.send_keys(self.email)
        
    async def wait_until_clickable(self, driver, xpath=None, class_name=None, duration=10000, frequency=0.01):
        if xpath:
            try: WebDriverWait(driver, duration, frequency).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            except: await asyncio.sleep(7)
        else:
            if class_name:
                try: WebDriverWait(driver, duration, frequency).until(EC.element_to_be_clickable((By.CLASS_NAME, class_name)))
                except: await asyncio.sleep(7)
            
    async def atk_request(self):
        link = self.driver.current_url
        link2 = link
        link = link.replace(r'%2F', '/')
        link = link[:-1]
        self.skuId = link[link.find('skuId=')+6 : link.find('&country')]
        self.checkoutId = link[link.find('checkoutId=')+11 : link.find('&launchId')]
        self.launchId = link[link.find('launchId=')+9 : link.find('&skuId')]
        self.itemName = link[link.rfind('/')+1 : ]

        phone = ''
        chars = 0
        for char in self.phone:
            if char.isdigit():
                phone += char
                chars += 1
                if chars == 3 or chars == 6 or chars == 8: phone += ' '

        index = self.index
        if len(index) == 6:
            index = index[:3] + '-' + index[3:]

        data = {"launchId":self.launchId,"checkoutId":self.checkoutId,"locale":"ru_RU","skuId":self.skuId,"shipping":{"address":{"address1":self.addressline,"address2":" ","address3":None,"city":self.city,"postalCode":index,"country":"RU","county":("RU-SPE" if self.city == "Санкт-Петербург" else "RU-MOW")},"recipient":{"firstName":self.name,"lastName":self.surname,"middleName":self.middlename,"email":self.email,"phoneNumber":"+7 " + phone},"method":"STANDARD"},"channel":"SNKRS","currency":"RUB","priceChecksum":self.price,"paymentToken":self.checkoutId}
        await asyncio.sleep(1)
        url2 = 'https://api.nike.com/launch/entries/v2'

        for cookies in self.driver.get_cookies(): 
            self.session.cookies.set(cookies['name'], cookies['value'])

        a = str(requests.utils.dict_from_cookiejar(self.session.cookies))
        a = a.replace(',', ';')
        a = a.replace(': ', '=')
        a = a.replace('{', '')
        a = a.replace('}', '')
        a = a.replace('\'', '')
        headers = {'Authorization': 'Bearer ' + str(self.auth),
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0", 
            'Referer': link2,
            'Cookie': a,
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
            'Content-Type' : 'application/json',
            'Host' : 'api.nike.com',
            'Origin' : 'https://gs.nike.com',
            'appid' : 'com.nike.commerce.snkrs.web',
            'Connection' : 'keep-alive'
        } 

        tasks = []
        xpath = '//button[@class="button-submit"]'
        self.place = True
        counter = 0
        until = float(datetime.now().replace(hour=int(self.hour), minute=int(self.minute), second=int(self.sec)).timestamp())
        sleep = int(until - float(time.time()))
        await asyncio.sleep(sleep)
                
        self.driver.find_element_by_xpath(xpath).click()

        now = time.time()
        sec = now + 5
        self.result = ""
    
        while now < sec and self.place:
            try:
                tasks.append(asyncio.create_task(self.send_request(headers, data, url2)))
            except: errors('in spam' + traceback.format_exc())

            await asyncio.sleep(0.1)
            now = time.time()
            counter += 1
        
        [await t for t in tasks] 
        
        errors(self.result, whichtask=self.whichtask)
        print(counter)
        await asyncio.sleep(420)

        response = self.session.get(url2, headers=headers)
        text = ""
        try:
            text = str(response.json())
        except: text = traceback.format_exc()
        with open(filePath + '/wins.txt', mode='a', encoding='utf-8') as f:
            f.write(text + '\n\n')
    
    async def send_request(self, headers, data, url):
        date = datetime.now()
        response = self.session.post(url, json=data, headers=headers)
        try:
            self.result += str(date) + " " + str(response.json()) + '\n'
            if 'code' in response.json() and response.json()['code'] == 'ENTRY_LIMIT_EXCEEDED': self.place = False
            if 'id' in response.json():
                self.place = False
                self.driver.get('https://www.nike.com/ru/launch/t/' + self.itemName + '/?launchEntryId=' + response.json()['id'] + '&launchId=' + response.json()['launchId'] + '&skuId=' + response.json()['skuId'])
        except: self.result += str(traceback.format_exc()) + '\n'
  
def errors(text, whichtask=1):
    with open(filePath + '/errors' + (str(whichtask) if whichtask != 1 else '') + '.txt', mode='a', encoding='utf-8') as f:
        try: f.write(text + '\n\n')
        except: f.write(traceback.format_exc())

async def start():
    with open(filePath + '/errors.txt', mode='w', encoding='utf-8') as f:
        f.write('')
    
    with open(filePath + '/wins.txt', mode='w', encoding='utf-8') as f:
        f.write('')

    wb = load_workbook(filePath + '/config.xlsx')
    sheet = wb['info']
    end = 8 + int(sheet.cell(row=2, column=4).value)

    tasks = []
    time_wait = sheet.cell(row=2, column=5).value
    hour = time_wait[:time_wait.find(':')]
    minutes = time_wait[time_wait.find(':')+1:]
    today = datetime.now()
    pause.until(datetime(int(today.year), int(today.month), int(today.day), int(hour), int(minutes), 0, 0))
    for whichtask in range(int(8), int(end)):
        try:
            with open(filePath + '/errors' + str(whichtask) + '.txt', mode='w', encoding='utf-8') as f:
                f.write('')

            n = nikeBot(whichtask, sheet)
            task = asyncio.create_task(n.main())
            tasks.append(task)
        except: errors('in start' + traceback.format_exc())
        
    [await t for t in tasks]

if __name__ == '__main__':
    html = requests.get('https://bilet.pp.ru/calculator_rus/tochnoe_moskovskoe_vremia.php').text
    pattern = r">(.*?)-(.*?)-(.*?)<"
    today = re.findall(pattern, html)[0]
    
    if int(today[2]) == 2021 and int(today[1]) <= 2:
        print("Успешный запуск программы")
        try: asyncio.run(start())
        except: print(traceback.format_exc())
    else: 
        print("Access denied")

    time.sleep(300)
    
#pyarmor pack -e " --onefile" --name "omrBotNike" --clean NikeBot.py